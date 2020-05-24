import time

import selenium
import openpyxl
import random
from lxml import etree
from Assistant import Assistant


class Spy:

    def __init__(self):
        self.assistant = Assistant()
        self.username = ""
        self.password = ""
        self.food_specific = ["Chinese"]
        self.max_num = 20
        self.urls = {
            "Hong kong": ["https://www.tripadvisor.com/Restaurants-g294217-Hong_Kong.html", "_1llCuDZj"]
        }
        self.information = {}

    def login(self, url: str) -> bool:
        self.assistant.chrome.get(url)
        self.assistant.chrome.implicitly_wait(5)
        self.assistant.chrome.find_element_by_link_text("Sign in").click()
        print("请在10秒内完成验证")
        self.assistant.chrome.implicitly_wait(10)
        username = self.assistant.chrome.find_element_by_id("regSignIn.email")
        password = self.assistant.chrome.find_element_by_id('regSignIn.password')
        username.send_keys(self.username)
        password.send_keys(self.password)

        self.assistant.chrome.find_element_by_class_name("ui_button primary regSubmitBtn").click()

        if not self.assistant.chrome.find_element_by_class_name("ui_button primary regSubmitBtn"):
            localtime = time.asctime(time.localtime(time.time()))
            print("{}:{} has logged in".format(localtime, self.username))
            return True
        else:
            localtime = time.asctime(time.localtime(time.time()))
            print("{}:Log Failed".format(localtime))
            return False

    def get_food_options(self) -> [[], {}]:
        start = time.time()
        print("Collecting Food Options:{}".format(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
        options = self.assistant.chrome.find_elements_by_class_name("option")
        temp = {}
        options_list = [a.text.split(' ') for a in options]
        for each in options_list:
            temp.setdefault(each[0], each[1].strip("()"))
        print("Collected Food Options. Time Used:{:.2f}".format(time.time() - start))
        return [options, temp]

    def user_info_analyse(self, info: str) -> {}:
        if '\n' in info:
            name, location = info.split('\n')
            if ', ' in location:
                city, country = location.split(', ')[-2], location.split(', ')[-1]
                return {
                    "Name": name,
                    "City": city,
                    "Country": country
                }
            else:
                return {
                    "Name": name,
                    "City": location,
                    "Country": "None"
                }
        else:
            return {
                "Name": info,
                "City": "None",
                "Country": "None"
            }

    def get_single_restaurant_detail_url(self, url: str, begin=0) -> {}:
        print("Grab Information from the {} page of {}".format(begin,url))
        if begin != 0:
            index = begin
            url = url.replace("Reviews", "Reviews-or{}".format(index * 10))

        self.assistant.chrome.get(url)
        time.sleep(8)

        resname = self.assistant.chrome.find_element_by_css_selector("div div h1").text
        if ', ' in resname:
            resname = resname.split(', ')[0]
        resname = resname.replace(" ", "_")

        print("开始爬取{}餐厅的评论信息".format(resname))
        start = time.time()

        html = self.assistant.chrome.execute_script("return document.documentElement.outerHTML")
        html = etree.HTML(html)
        count = html.xpath("//a[@class='pageNum last   cx_brand_refresh_phase2']")[0].text
        count = int(count.strip("()"))

        if begin == 0:
            index = 0
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "{}".format(resname)
            ws.append(["User Name", "City", "Country", "Date", "Comment"])
        else:
            index = begin
            wb = openpyxl.load_workbook("{}.xlsx".format(resname))
            ws = wb.active
        count_chinese=1
        while index < count:
            index += 1
            print("抓取第{}页的评论信息".format(index))
            try:
                # Expansion
                self.assistant.chrome.find_element_by_xpath("//span[@class='taLnk ulBlueLinks']").click()
                time.sleep(3)

                reviews = self.assistant.chrome.find_elements_by_xpath("//div[@class='review-container']")
                if len(reviews)<10 and index!=count:
                    print("抓取评论数量不完整")
                    self.assistant.chrome.refresh()
                    time.sleep(10)
                    reviews = self.assistant.chrome.find_elements_by_xpath("//div[@class='review-container']")
                    print("现在抓取了{}个评论".format(len(reviews)))
                user = self.assistant.chrome.find_elements_by_xpath(
                    "//div[@class='info_text pointer_cursor' or (@class='info_text ')]")
                data = reviews[0].find_elements_by_xpath("//span[@class='ratingDate']")
            except Exception:
                print("再等待10秒")
                time.sleep(10)
                try:
                    # Expansion
                    self.assistant.chrome.find_element_by_xpath("//span[@class='taLnk ulBlueLinks']").click()
                    time.sleep(3)
                except selenium.common.exceptions.NoSuchElementException:
                    reviews = self.assistant.chrome.find_elements_by_xpath("//div[@class='review-container']")
                    user = self.assistant.chrome.find_elements_by_xpath("//div[@class='info_text pointer_cursor' or (@class='info_text ')]")
                    data = reviews[0].find_elements_by_xpath("//span[@class='ratingDate']")
            c = 0
            for review in reviews:
                comment = review.find_element_by_class_name("partial_entry").text
                user_info = self.user_info_analyse(user[c].text)
                if user_info["Country"] == "China":
                    print("\t\tFind Chinese\t\tUser:{}\t\tCity:{}\tCountry:{}\tDate:{}\tCount:{}".format(user_info['Name'],
                                                                                                   user_info['City'],
                                                                                                   user_info['Country'],
                                                                                                   data[c].text.strip(
                                                                                                       "Reviewed "),
                                                                                                   count_chinese
                                                                                                   ))
                    count_chinese+=1
                    c += 1
                    continue
                else:
                    print("\t\tUser:{}\t\tCity:{}\tCountry:{}\tDate:{}".format(user_info['Name'],
                                                                             user_info['City'],
                                                                             user_info['Country'],
                                                                             data[c].text.strip("Reviewed ")))
                    ws.append(
                        [user_info['Name'], user_info['City'], user_info['Country'], data[c].text.strip("Reviewed "),
                         comment])
                    if index % 3 == 0:
                        wb.save("{}.xlsx".format(resname))
                        wb = openpyxl.load_workbook("{}.xlsx".format(resname))
                        ws = wb.active
                    c += 1
            if begin == 0:
                print(url.replace("Reviews", "Reviews-or{}".format(index * 10)))
                self.assistant.chrome.get(url.replace("Reviews", "Reviews-or{}".format(index * 10)))
            else:
                url = url.replace("Reviews-or{}".format(index * 10 - 10), "Reviews-or{}".format(index * 10))
                print(url)
                self.assistant.chrome.get(url)
            time.sleep(random.randint(5, 12))
        wb.save("{}.xlsx".format(resname))
        print("\t{}餐厅的评论信息爬取完成. 用时:{:.2f}".format(resname, time.time() - start))

    # def get_single_restaurant_detail(self, restaurant: object) -> {}:
    #     temp = {}
    #     start = time.time()
    #     self.assistant.chrome.implicitly_wait(10)
    #     restaurant.find_element_by_css_selector("div div div span a").click()
    #     handles = self.assistant.chrome.window_handles
    #     # Switch to the new window
    #     self.assistant.chrome.switch_to.window(handles[1])
    #
    #     name = self.assistant.chrome.find_element_by_css_selector("div div h1").text
    #     print("\t开始爬取{}餐厅的评论信息".format(name))
    #     count = self.assistant.chrome.find_element_by_xpath("//span[@class='reviews_header_count']").text
    #     count = int(count.strip("()"))
    #     temp.setdefault(name, {})
    #
    #     # Expansion
    #     self.assistant.chrome.find_element_by_xpath("//span[@class='taLnk ulBlueLinks']").click()
    #     time.sleep(3)
    #     reviews = self.assistant.chrome.find_elements_by_xpath("//div[@class='review-container']")
    #
    #     index = 0
    #     while index < count:
    #         for review in reviews:
    #             index += 1
    #             user = review.find_element_by_xpath("//div[@class='info_text pointer_cursor']").text
    #             comment = review.find_element_by_class_name("partial_entry").text
    #             data = review.find_element_by_xpath("//div[@class='prw_rup prw_reviews_stay_date_hsx']").text
    #             data = data.split(": ")[1]
    #             temp[name].setdefault(index, {"User": user, "Comment": comment, "Data": data})
    #             print("评论{}:{}".format(index, comment))
    #         try:
    #             next = spy.assistant.chrome.find_element_by_xpath(
    #                 "//a[@class='nav next ui_button primary  cx_brand_refresh_phase2']")
    #             next.click()
    #             self.assistant.chrome.implicitly_wait(5)
    #         except selenium.common.exceptions.NoSuchElementException:
    #             print("\t餐厅{}全部评论爬取完毕，共{}条，耗时{:.2f}".format(name, count, time.time() - start))
    #             self.assistant.chrome.close()
    #             return temp
    #     self.assistant.chrome.close()
    #     return temp

    def get_all_restaurant_url(self, url: str, type: str) -> [str]:
        self.assistant.chrome.get(url)
        time.sleep(3)
        option_list, option_info = self.get_food_options()
        urls = []
        for each in option_list:
            text = each.text.split(" ")[0]
            if text == type:
                each.click()
                time.sleep(4)
                restaurants = self.assistant.chrome.find_elements_by_xpath("//div[@class='wQjYiB7z']/span/a")
                urls = [res.get_property('href') for res in restaurants]
                print("共找到{}个餐厅，他们的地址如下:".format(len(urls)))
                print(urls)
                break
            else:
                print("没有找到所需要类型的餐厅")
        return urls

    def get_all_restaurant_detail(self, city: str, option_list: []) -> {}:
        temp = {}
        temp.setdefault(city, {})
        for each in option_list:
            text = each.text.split(" ")[0]
            if text in self.food_specific:
                each.click()
                restaurants = self.assistant.chrome.find_elements_by_class_name(self.urls[city][1])
                print("Collecting Data of {}.".format(text))
                count = 0
                for res in restaurants:
                    if count > self.max_num:
                        return temp
                    self.assistant.chrome.implicitly_wait(2)
                    review = self.get_single_restaurant_detail(res)
                    temp[city].update(review)
                    count += 1
        return temp

    def get_information(self, city: str):
        start = time.time()
        print("Collecting Informations for {}:{}".format(city, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
        if self.urls.__contains__(city):
            self.assistant.chrome.get(self.urls[city][0])
            self.assistant.chrome.implicitly_wait(2)
            temp = {}

            # For options
            options_list, option_dict = self.get_food_options()

            # For a Specific Food
            info = self.get_all_restaurant_detail(city, options_list)
            print("Collecting Informations for {} Done. Time Used:{:.2f}".format(city, time.time() - start))
        else:
            print("Input City doesn't exit!")


if __name__ == '__main__':
    spy = Spy()
    # urls=spy.get_all_restaurant_url("https://www.tripadvisor.com/Restaurants-g294217-Hong_Kong.html","Chinese")
    urls = ['https://www.tripadvisor.com/Restaurant_Review-g294217-d1792423-Reviews-Man_Wah-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2399904-Reviews-Tin_Lung_Heen-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d927163-Reviews-Yan_Toh_Heen_InterContinental_Hong_Kong-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d797531-Reviews-Dynasty-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d10712103-Reviews-Cloudland_Chinese_Cuisine-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d816180-Reviews-Celestial_Court_Chinese_Restaurant-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2095035-Reviews-Sing_Yin_Cantonese_Dining-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d1062729-Reviews-Lung_King_Heen-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2540306-Reviews-One_Dim_Sum_Chinese_Restaurant-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2461243-Reviews-The_Chairman-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2184677-Reviews-One_Harbour_Road_Grand_Hyatt_Hong_Kong-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d7693632-Reviews-China_Club-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d3330314-Reviews-Islamic_Centre_Canteen-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2540127-Reviews-The_Monogamous_Chinese-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2706744-Reviews-Din_Tai_Fung_Yee_Wo_Branch-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d11954852-Reviews-Restaurant_de_Chine-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d5400408-Reviews-Din_Tai_Fung_Silvercord-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d6579151-Reviews-Mott_32-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d812337-Reviews-Spring_Moon-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2428744-Reviews-Shang_Palace-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d6739503-Reviews-Ding_Dim_1968-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2390123-Reviews-Above_and_Beyond-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d3749262-Reviews-Tim_Ho_Wan_Sham_Shui_Po-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d1367659-Reviews-Hutong-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2364439-Reviews-Man_Ho_Chinese_Restaurant-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d15128192-Reviews-The_Chinese_Library-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d7178801-Reviews-Din_Tai_Fung_Miramar-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d4161889-Reviews-Tsim_Chai_Kee-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2437900-Reviews-DimDimSum_Dim_Sum_Specialty_Store_Jordan-Hong_Kong.html',
            'https://www.tripadvisor.com/Restaurant_Review-g294217-d2534180-Reviews-Dim_Sum_Square-Hong_Kong.html']
    for url in urls[0:]:
        spy.get_single_restaurant_detail_url(url)
    spy.assistant.chrome.close()
    # spy.get_single_restaurant_detail_url(
    #     url="https://www.tripadvisor.com/Restaurant_Review-g294217-d1792423-Reviews-Man_Wah-Hong_Kong.html",begin=41)
